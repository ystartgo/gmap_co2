from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage

HEADERS = [
    "查查屬區",
    "申請單號",
    "員工編號",
    "員工姓名",
    "出發日期",
    "起點",
    "終點",
    "交通工具",
    "乘客數",
    "每人單次里程(pkm)",
    "運輸碳排放係數(kgCO2e/pkm)",
    "CO2e(kg)",
]

def _get_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

def _find_map_anchor(wb):
    candidates = ["MAP_AREA", "截圖區", "MAP_SHOT", "地圖截圖"]
    for n in candidates:
        dn = wb.defined_names.get(n)
        if dn is None:
            continue
        dests = dn.destinations
        for title, coord in dests:
            sheet = wb[title]
            cell = coord.split("!")[-1]
            if ":" in cell:
                cell = cell.split(":")[0]
            return sheet, cell
    return None, None

def _init_headers(ws):
    if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=i).value = h

def _find_col(ws, header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    return None

def write_row_and_image(
    xlsx: str,
    dept: str,
    ticket: str,
    emp_id: str,
    emp_name: str,
    date,
    origin: str,
    destination: str,
    mode: str,
    passengers: int,
    distance_km: float,
    factor: float,
    co2e_kg: float,
    image_path: str,
):
    p = Path(xlsx)
    if p.exists():
        wb = load_workbook(str(p))
    else:
        wb = Workbook()
    ws = _get_sheet(wb, "行程明細")
    _init_headers(ws)
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = dept
    ws.cell(row=row, column=2).value = ticket
    ws.cell(row=row, column=3).value = emp_id
    ws.cell(row=row, column=4).value = emp_name
    ws.cell(row=row, column=5).value = date.strftime("%Y-%m-%d")
    ws.cell(row=row, column=6).value = origin
    ws.cell(row=row, column=7).value = destination
    ws.cell(row=row, column=8).value = mode
    col_pass = _find_col(ws, "乘客數") or 9
    col_dist = _find_col(ws, "每人單次里程(pkm)") or 10
    col_factor = _find_col(ws, "運輸碳排放係數(kgCO2e/pkm)") or 11
    col_co2 = _find_col(ws, "CO2e(kg)") or 12
    ws.cell(row=row, column=col_pass).value = passengers
    ws.cell(row=row, column=col_dist).value = distance_km
    ws.cell(row=row, column=col_factor).value = factor
    ws.cell(row=row, column=col_co2).value = co2e_kg
    imgs = _get_sheet(wb, "截圖")
    anchor_row = imgs.max_row + 1 if imgs.max_row else 1
    img = XLImage(image_path)
    ms, mc = _find_map_anchor(wb)
    if ms is not None and mc is not None:
        ms.add_image(img, mc)
    else:
        imgs.add_image(img, f"A{anchor_row}")
    wb.save(str(p))

def scan_and_fill(xlsx: str, outdir: str):
    p = Path(xlsx)
    if not p.exists():
        return
    wb = load_workbook(str(p))
    ws = _get_sheet(wb, "行程明細")
    if ws.max_row <= 1:
        wb.save(str(p))
        return
    from .gmaps import capture_route
    from .factors import load_factors, compute
    factors = load_factors()
    col_mode = _find_col(ws, "交通工具") or 8
    col_origin = _find_col(ws, "起點") or 6
    col_dest = _find_col(ws, "終點") or 7
    col_pass = _find_col(ws, "乘客數") or 9
    col_dist = _find_col(ws, "每人單次里程(pkm)") or 10
    col_factor = _find_col(ws, "運輸碳排放係數(kgCO2e/pkm)") or 11
    col_co2 = _find_col(ws, "CO2e(kg)") or 12
    for row in range(2, ws.max_row + 1):
        mode = ws.cell(row=row, column=col_mode).value
        origin = ws.cell(row=row, column=col_origin).value
        destination = ws.cell(row=row, column=col_dest).value
        passengers = ws.cell(row=row, column=col_pass).value
        if isinstance(passengers, (int, float)) and passengers > 0:
            pass_count = int(passengers)
        else:
            pass_count = 1
        shot_path, dist = capture_route(origin, destination, mode, outdir, None)
        factor = factors.get(mode)
        pkm = (dist or 0.0) * pass_count
        co2 = compute(pkm, factor)
        ws.cell(row=row, column=col_dist).value = dist
        ws.cell(row=row, column=col_factor).value = factor
        ws.cell(row=row, column=col_co2).value = co2
        imgs = _get_sheet(wb, "截圖")
        anchor_row = imgs.max_row + 1 if imgs.max_row else 1
        img = XLImage(shot_path)
        imgs.add_image(img, f"A{anchor_row}")
    wb.save(str(p))
